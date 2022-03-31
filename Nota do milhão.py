#!/usr/bin/env python
# coding: utf-8

# In[7]:


from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import Select


#nesse bloco caminho vai receber o endereço da planilha e preparar ela para o laço de repetição
caminho = 'teste nota do milhão.xlsx'
arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active
max_linha = planilha1.max_row
max_coluna = planilha1.max_column
for i in range(2, max_linha + 1):
    for j in range(1, max_coluna + 1):
        #Nesse bloco resultado recebe linhas e colunas da planilha e preeche os campos de login
        resultado = planilha1.cell(row=i, column=4).value
        CNPJ = resultado
        #Abre a pagina solicitado
        driver = webdriver.Chrome()
        driver.get('https://nfe.prefeitura.sp.gov.br/login.aspx')
       
        #seleciona e preeche o campo login
        driver.find_element_by_xpath('//*[@id="btnAtualizar"]/img').click()
        time.sleep(1)
        campo_login = driver.find_element_by_xpath('//*[@id="ctl00_body_tbCpfCnpj"]')
        campo_login.send_keys(resultado)
        time.sleep(2)
       
        #seleciona e preeche o campo senha
        resultado = planilha1.cell(row=i, column=5).value
        campo_senha = driver.find_element_by_xpath('//*[@id="ctl00_body_tbSenha"]')
        campo_senha.send_keys(resultado)
        time.sleep(10)
        driver.find_element_by_xpath('//*[@id="ctl00_body_btEntrar"]').click()
        time.sleep(2)
        
        #Notas
        driver.find_element_by_xpath('//*[@id="ctl00_wpMenuLateral_mnuRotinasn3"]/td/table/tbody/tr/td/a').click()
        driver.find_element_by_xpath('//*[@id="ctl00_body_ddlPrestador"]').click()
        resultado = planilha1.cell(row=i, column=2).value
        print(resultado)
        # Indicando quem é o Prestador
        segunda_tela = driver.find_element_by_xpath('//*[@id="ctl00_body_tbCPFCNPJTomador"]')
        segunda_tela.send_keys(CNPJ)
        select = Select(driver.find_element_by_id('ctl00_body_ddlPrestador'))
        select.select_by_visible_text(resultado)
        # Avançar a pagina
        driver.find_element_by_xpath('//*[@id="ctl00_body_btAvancar"]').click()
        # Valor da nota
        resultado = planilha1.cell(row=i, column=3).value
        valor_pg4 = driver.find_element_by_xpath('//*[@id="ctl00_body_tbValor"]')
        valor_pg4.send_keys(resultado)
        #driver.find_element_by_xpath('//*[@id="ctl00_body_btEmitir"]').click()
        
        
      
        

